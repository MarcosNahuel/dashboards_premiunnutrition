import os
import glob
import json
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = r'D:\OneDrive\GitHub\premiumnutrition'
EXEC_DIR = os.path.join(BASE_DIR, 'EJECUCIONES')
PATTERN = os.path.join(EXEC_DIR, 'ejecucion-*-success.json')
EXPECTED_CSV = os.path.join(BASE_DIR, 'dashboards_premiunnutrition', 'analysis_outputs', 'tabla_respuestas.csv')

expected_map = {}
with open(EXPECTED_CSV, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        try:
            expected_map[int(row['Id'])] = row.get('ExpectedResponse', '')
        except (ValueError, KeyError):
            continue

def extract_plain_response(text: str) -> str:
    if not text:
        return ''
    match = re.search(r"`json\s*(\{.*?\})\s*`", text, re.DOTALL)
    if match:
        payload = match.group(1)
        try:
            parsed = json.loads(payload)
        except json.JSONDecodeError:
            pass
        else:
            respuesta = parsed.get('output', {}).get('respuesta')
            if isinstance(respuesta, list):
                return '\n'.join(str(item) for item in respuesta)
            if isinstance(respuesta, str):
                return respuesta
    return text

rows = []
for filepath in sorted(glob.glob(PATTERN), key=lambda p: int(os.path.basename(p).split('-')[1])):
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    run_data = data.get('data', {}).get('resultData', {}).get('runData', {})
    try:
        question = run_data['Webhook1'][0]['data']['main'][0][0]['json']['body']['message[add][0][text]']
    except (KeyError, IndexError, TypeError):
        continue
    node_runs = run_data.get('OpenAI Chat Model2')
    if not node_runs:
        continue
    actual_raw = None
    for run in node_runs:
        try:
            gens = run['data']['ai_languageModel'][0][0]['json']['response']['generations']
            if gens and gens[0] and gens[0][0].get('text'):
                actual_raw = gens[0][0]['text']
                break
        except (KeyError, IndexError, TypeError):
            continue
    if actual_raw is None:
        continue
    actual_clean = extract_plain_response(actual_raw)
    run_id = int(data['id'])
    rows.append((run_id, question, actual_clean, expected_map.get(run_id, '')))

wb = Workbook()
ws = wb.active
ws.title = 'Respuestas'
ws.append(['Id', 'Pregunta', 'Respuesta agente', 'Respuesta esperada'])
for row in rows:
    ws.append(row)

wrap = Alignment(wrap_text=True, vertical='top')
col_widths = {1: 8, 2: 60, 3: 90, 4: 90}
for col in range(1, ws.max_column + 1):
    letter = get_column_letter(col)
    ws.column_dimensions[letter].width = col_widths.get(col, 25)
    for row_idx in range(1, ws.max_row + 1):
        ws.cell(row=row_idx, column=col).alignment = wrap

OUT_PATH = os.path.join(BASE_DIR, 'dashboards_premiunnutrition', 'analysis_outputs', 'tabla_respuestas_final.xlsx')
wb.save(OUT_PATH)
print(f'Saved: {OUT_PATH}')
print(f'Rows: {len(rows)}')
