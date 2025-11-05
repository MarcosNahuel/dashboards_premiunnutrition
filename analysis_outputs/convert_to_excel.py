import sys
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

if len(sys.argv) != 3:
    raise SystemExit("Usage: python convert_to_excel.py <src_csv> <dst_xlsx>")

src_path, dst_path = sys.argv[1:3]

wb = Workbook()
ws = wb.active
ws.title = "Respuestas"

with open(src_path, "r", encoding="utf-8") as f:
    reader = csv.reader(f)
    headers = next(reader)
    ws.append(headers)
    for row in reader:
        ws.append(row)

wrap = Alignment(wrap_text=True, vertical="top")
for col in range(1, ws.max_column + 1):
    letter = get_column_letter(col)
    ws.column_dimensions[letter].width = 45 if col > 1 else 8
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = wrap

wb.save(dst_path)
